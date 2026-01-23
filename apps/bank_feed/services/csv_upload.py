"""
CSV/Excel file parsing service for bank transaction uploads.
Handles file parsing, date/amount detection, and category matching.
"""

import csv
import io
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

from dateutil import parser as date_parser
from django.db.models import Q
from openpyxl import load_workbook

from apps.accounts.models import Account


@dataclass
class ParsedTransaction:
    """Represents a parsed transaction from CSV/Excel."""

    row_number: int
    date: date | None
    description: str | None
    payee: str | None
    category: str | None
    amount: Decimal | None
    error: str | None = None
    matched_category_id: int | None = None
    is_potential_duplicate: bool = False


@dataclass
class ParseResult:
    """Result of parsing a file."""

    headers: list[str]
    sample_rows: list[list[str]]
    total_rows: int
    error: str | None = None


@dataclass
class PreviewResult:
    """Result of previewing parsed transactions."""

    transactions: list[ParsedTransaction]
    unmapped_categories: list[str]
    error_count: int
    duplicate_count: int


# Common date formats to try before falling back to dateutil
DATE_FORMATS = [
    "%Y-%m-%d",  # 2024-01-15
    "%m/%d/%Y",  # 01/15/2024
    "%d/%m/%Y",  # 15/01/2024
    "%m-%d-%Y",  # 01-15-2024
    "%d-%m-%Y",  # 15-01-2024
    "%Y/%m/%d",  # 2024/01/15
    "%m/%d/%y",  # 01/15/24
    "%d/%m/%y",  # 15/01/24
    "%b %d, %Y",  # Jan 15, 2024
    "%B %d, %Y",  # January 15, 2024
    "%d %b %Y",  # 15 Jan 2024
    "%d %B %Y",  # 15 January 2024
]


def detect_encoding(file_content: bytes) -> str:
    """Detect the encoding of file content."""
    # Try common encodings
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252", "iso-8859-1"]

    for encoding in encodings:
        try:
            file_content.decode(encoding)
            return encoding
        except (UnicodeDecodeError, LookupError):
            continue

    return "utf-8"  # Default fallback


def parse_date(value: str) -> date | None:
    """
    Parse a date string using common formats, falling back to dateutil.
    Returns None if parsing fails.
    """
    if not value or not value.strip():
        return None

    value = value.strip()

    # Try common formats first (faster)
    from datetime import datetime

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    # Fall back to dateutil for more flexible parsing
    try:
        return date_parser.parse(value, dayfirst=False).date()
    except (ValueError, TypeError):
        pass

    # Try with dayfirst=True for ambiguous dates like 01/02/2024
    try:
        return date_parser.parse(value, dayfirst=True).date()
    except (ValueError, TypeError):
        pass

    return None


def parse_amount(value: str) -> Decimal | None:
    """
    Parse an amount string, handling:
    - Currency symbols ($, €, £, etc.)
    - Thousands separators (comma or period)
    - Accounting format (parentheses for negatives)
    - Negative signs

    Returns None if parsing fails.
    """
    if not value or not value.strip():
        return None

    value = value.strip()

    # Check for accounting format (parentheses = negative)
    is_negative = False
    if value.startswith("(") and value.endswith(")"):
        is_negative = True
        value = value[1:-1]

    # Check for explicit negative sign
    if value.startswith("-"):
        is_negative = True
        value = value[1:]

    # Remove currency symbols and whitespace
    value = re.sub(r"[$€£¥₹\s]", "", value)

    # Handle thousands separators
    # If we have both commas and periods, determine which is the decimal separator
    has_comma = "," in value
    has_period = "." in value

    if has_comma and has_period:
        # Find positions - the one that appears last is likely the decimal separator
        last_comma = value.rfind(",")
        last_period = value.rfind(".")

        if last_period > last_comma:
            # Period is decimal separator, remove commas
            value = value.replace(",", "")
        else:
            # Comma is decimal separator, remove periods and convert comma
            value = value.replace(".", "").replace(",", ".")
    elif has_comma:
        # Could be thousands separator (1,000) or decimal separator (1,50)
        # Check if there are exactly 2 digits after the comma (European decimal)
        parts = value.split(",")
        if len(parts) == 2 and len(parts[1]) == 2:
            value = value.replace(",", ".")
        else:
            value = value.replace(",", "")

    try:
        amount = Decimal(value)
        if is_negative:
            amount = -amount
        return amount
    except InvalidOperation:
        return None


def parse_csv_file(file: BinaryIO) -> ParseResult:
    """
    Parse a CSV file and return headers, sample rows, and total count.
    """
    try:
        content = file.read()
        encoding = detect_encoding(content)
        text = content.decode(encoding)

        # Use csv.Sniffer to detect dialect
        try:
            dialect = csv.Sniffer().sniff(text[:8192])
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        rows = list(reader)

        if not rows:
            return ParseResult(headers=[], sample_rows=[], total_rows=0, error="File is empty")

        headers = rows[0]
        data_rows = rows[1:]

        # Get sample rows (up to 5)
        sample_rows = data_rows[:5]

        return ParseResult(
            headers=headers,
            sample_rows=sample_rows,
            total_rows=len(data_rows),
        )

    except Exception as e:
        return ParseResult(headers=[], sample_rows=[], total_rows=0, error=str(e))


def parse_excel_file(file: BinaryIO) -> ParseResult:
    """
    Parse an Excel file and return headers, sample rows, and total count.
    """
    try:
        workbook = load_workbook(filename=file, read_only=True, data_only=True)
        sheet = workbook.active

        rows = []
        for row in sheet.iter_rows(values_only=True):
            # Convert all values to strings
            str_row = [str(cell) if cell is not None else "" for cell in row]
            # Skip completely empty rows
            if any(cell.strip() for cell in str_row):
                rows.append(str_row)

        workbook.close()

        if not rows:
            return ParseResult(headers=[], sample_rows=[], total_rows=0, error="File is empty")

        headers = rows[0]
        data_rows = rows[1:]

        # Get sample rows (up to 5)
        sample_rows = data_rows[:5]

        return ParseResult(
            headers=headers,
            sample_rows=sample_rows,
            total_rows=len(data_rows),
        )

    except Exception as e:
        return ParseResult(headers=[], sample_rows=[], total_rows=0, error=str(e))


def parse_file(file: BinaryIO, filename: str) -> ParseResult:
    """
    Parse a file based on its extension.
    """
    filename_lower = filename.lower()

    if filename_lower.endswith((".xlsx", ".xls")):
        return parse_excel_file(file)
    elif filename_lower.endswith(".csv"):
        return parse_csv_file(file)
    else:
        return ParseResult(
            headers=[],
            sample_rows=[],
            total_rows=0,
            error="Unsupported file type. Please upload a CSV or Excel file.",
        )


def match_category(category_name: str, team) -> Account | None:
    """
    Match a category name to an existing Account.
    Matches by name (case-insensitive) or account_number.
    """
    if not category_name or not category_name.strip():
        return None

    category_name = category_name.strip()

    # Try exact match on name (case-insensitive) or account_number
    account = Account.objects.filter(team=team).filter(
        Q(name__iexact=category_name) | Q(account_number__iexact=category_name)
    ).first()

    return account


def get_all_rows_from_csv(file: BinaryIO) -> list[list[str]]:
    """Read all data rows from a CSV file."""
    content = file.read()
    encoding = detect_encoding(content)
    text = content.decode(encoding)

    try:
        dialect = csv.Sniffer().sniff(text[:8192])
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)

    if not rows:
        return []

    return rows[1:]  # Skip header row


def get_all_rows_from_excel(file: BinaryIO) -> list[list[str]]:
    """Read all data rows from an Excel file."""
    workbook = load_workbook(filename=file, read_only=True, data_only=True)
    sheet = workbook.active

    rows = []
    for row in sheet.iter_rows(values_only=True):
        str_row = [str(cell) if cell is not None else "" for cell in row]
        if any(cell.strip() for cell in str_row):
            rows.append(str_row)

    workbook.close()

    if not rows:
        return []

    return rows[1:]  # Skip header row


def preview_transactions(
    file: BinaryIO,
    filename: str,
    column_mapping: dict,
    category_mappings: dict,
    team,
    account_id: int,
) -> PreviewResult:
    """
    Parse file with column mapping and return preview of transactions.

    Args:
        file: The uploaded file
        filename: Original filename (for type detection)
        column_mapping: Dict mapping our fields to column indices
            {
                "date": 0,
                "description": 1,
                "payee": 2,  # optional
                "category": 3,  # optional
                "amount": 4,  # for single amount column
                "inflow": 4,  # for dual amount columns
                "outflow": 5,  # for dual amount columns
            }
        category_mappings: Dict mapping category names to account IDs
        team: The team object
        account_id: The bank account ID being uploaded to

    Returns:
        PreviewResult with parsed transactions
    """
    from apps.bank_feed.models import BankTransaction

    filename_lower = filename.lower()

    if filename_lower.endswith((".xlsx", ".xls")):
        rows = get_all_rows_from_excel(file)
    elif filename_lower.endswith(".csv"):
        rows = get_all_rows_from_csv(file)
    else:
        return PreviewResult(
            transactions=[],
            unmapped_categories=[],
            error_count=1,
            duplicate_count=0,
        )

    transactions = []
    unmapped_categories_set = set()
    error_count = 0
    duplicate_count = 0

    # Get column indices from mapping
    date_col = column_mapping.get("date")
    desc_col = column_mapping.get("description")
    payee_col = column_mapping.get("payee")
    category_col = column_mapping.get("category")
    amount_col = column_mapping.get("amount")
    inflow_col = column_mapping.get("inflow")
    outflow_col = column_mapping.get("outflow")

    use_dual_amount = inflow_col is not None or outflow_col is not None

    for row_num, row in enumerate(rows, start=2):  # Start at 2 (1-indexed, skip header)
        error = None
        parsed_date = None
        description = None
        payee = None
        category_name = None
        amount = None
        matched_category_id = None

        # Parse date
        if date_col is not None and date_col < len(row):
            parsed_date = parse_date(row[date_col])
            if not parsed_date:
                error = f"Invalid date: {row[date_col]}"

        # Parse description
        if desc_col is not None and desc_col < len(row):
            description = row[desc_col].strip() if row[desc_col] else None

        # Parse payee
        if payee_col is not None and payee_col < len(row):
            payee = row[payee_col].strip() if row[payee_col] else None

        # Parse category
        if category_col is not None and category_col < len(row):
            category_name = row[category_col].strip() if row[category_col] else None

        # Parse amount
        if use_dual_amount:
            # Dual column mode: inflow becomes negative, outflow becomes positive
            inflow_val = Decimal("0")
            outflow_val = Decimal("0")

            if inflow_col is not None and inflow_col < len(row) and row[inflow_col].strip():
                inflow_val = parse_amount(row[inflow_col])
                if inflow_val is None:
                    error = f"Invalid inflow amount: {row[inflow_col]}"
                    inflow_val = Decimal("0")
                else:
                    inflow_val = abs(inflow_val)

            if outflow_col is not None and outflow_col < len(row) and row[outflow_col].strip():
                outflow_val = parse_amount(row[outflow_col])
                if outflow_val is None:
                    error = f"Invalid outflow amount: {row[outflow_col]}"
                    outflow_val = Decimal("0")
                else:
                    outflow_val = abs(outflow_val)

            # Convert to Plaid convention: positive = outflow, negative = inflow
            amount = outflow_val - inflow_val
        else:
            # Single column mode: use value as-is (positive = expense, negative = income)
            if amount_col is not None and amount_col < len(row):
                amount = parse_amount(row[amount_col])
                if amount is None and row[amount_col].strip():
                    error = f"Invalid amount: {row[amount_col]}"

        # Match category
        if category_name:
            # First check if user has mapped this category
            if category_name in category_mappings:
                matched_category_id = category_mappings[category_name]
            else:
                # Try to auto-match
                matched_account = match_category(category_name, team)
                if matched_account:
                    matched_category_id = matched_account.id
                else:
                    unmapped_categories_set.add(category_name)

        # Check for potential duplicates
        is_duplicate = False
        if parsed_date and amount is not None:
            existing = BankTransaction.objects.filter(
                team=team,
                account_id=account_id,
                posted_date=parsed_date,
                amount=amount,
            )
            if description:
                existing = existing.filter(description__iexact=description)
            if existing.exists():
                is_duplicate = True
                duplicate_count += 1

        if error:
            error_count += 1

        transactions.append(
            ParsedTransaction(
                row_number=row_num,
                date=parsed_date,
                description=description,
                payee=payee,
                category=category_name,
                amount=amount,
                error=error,
                matched_category_id=matched_category_id,
                is_potential_duplicate=is_duplicate,
            )
        )

    return PreviewResult(
        transactions=transactions,
        unmapped_categories=list(unmapped_categories_set),
        error_count=error_count,
        duplicate_count=duplicate_count,
    )


def create_transactions(
    transactions: list[dict],
    team,
    account_id: int,
    skip_duplicates: bool = True,
) -> dict:
    """
    Create BankTransaction records from parsed transactions.

    Args:
        transactions: List of transaction dicts with keys:
            date, description, payee, category_id, amount, skip (optional)
        team: The team object
        account_id: The bank account ID
        skip_duplicates: Whether to skip potential duplicates

    Returns:
        Dict with created_count, skipped_count, and error_count
    """
    from apps.bank_feed.models import BankTransaction

    created_count = 0
    skipped_count = 0
    error_count = 0

    for tx_data in transactions:
        # Skip if marked to skip
        if tx_data.get("skip"):
            skipped_count += 1
            continue

        # Validate required fields
        if not tx_data.get("date") or tx_data.get("amount") is None:
            error_count += 1
            continue

        # Check for duplicates if skip_duplicates is True
        if skip_duplicates:
            existing = BankTransaction.objects.filter(
                team=team,
                account_id=account_id,
                posted_date=tx_data["date"],
                amount=tx_data["amount"],
            )
            if tx_data.get("description"):
                existing = existing.filter(description__iexact=tx_data["description"])
            if existing.exists():
                skipped_count += 1
                continue

        # Create the BankTransaction
        bank_tx = BankTransaction.objects.create(
            team=team,
            account_id=account_id,
            posted_date=tx_data["date"],
            description=tx_data.get("description") or "",
            merchant_name=tx_data.get("payee"),
            amount=tx_data["amount"],
            source=BankTransaction.SOURCE_CSV,
            raw=tx_data.get("raw"),
        )

        created_count += 1

        # Auto-categorize if category_id is provided
        if tx_data.get("category_id"):
            _auto_categorize_transaction(bank_tx, tx_data["category_id"], team)

    return {
        "created_count": created_count,
        "skipped_count": skipped_count,
        "error_count": error_count,
    }


def _auto_categorize_transaction(bank_tx, category_id: int, team):
    """
    Auto-categorize a bank transaction by creating a journal entry.
    Reuses the logic from BankFeedViewSet._create_journal_from_bank_transaction.
    """
    from apps.accounts.models import Account
    from apps.journal.models import JournalEntry, JournalLine

    try:
        category_account = Account.objects.get(id=category_id, team=team)
    except Account.DoesNotExist:
        return  # Skip if category doesn't exist

    # Create journal entry
    journal_entry = JournalEntry.objects.create(
        team=team,
        entry_date=bank_tx.posted_date,
        description=bank_tx.description,
        source=bank_tx.source,
        status=JournalEntry.STATUS_POSTED,
    )

    # Calculate amounts (Plaid convention: positive = outflow, negative = inflow)
    amount = abs(bank_tx.amount)
    is_inflow = bank_tx.amount < 0

    # Create journal lines
    if is_inflow:
        # Money coming in: debit bank account, credit category
        JournalLine.objects.create(
            journal_entry=journal_entry,
            team=team,
            account=bank_tx.account,
            dr_amount=amount,
            cr_amount=Decimal("0"),
        )
        JournalLine.objects.create(
            journal_entry=journal_entry,
            team=team,
            account=category_account,
            dr_amount=Decimal("0"),
            cr_amount=amount,
        )
    else:
        # Money going out: credit bank account, debit category
        JournalLine.objects.create(
            journal_entry=journal_entry,
            team=team,
            account=bank_tx.account,
            dr_amount=Decimal("0"),
            cr_amount=amount,
        )
        JournalLine.objects.create(
            journal_entry=journal_entry,
            team=team,
            account=category_account,
            dr_amount=amount,
            cr_amount=Decimal("0"),
        )

    # Link the bank transaction to the journal entry
    bank_tx.journal_entry = journal_entry
    bank_tx.save()
